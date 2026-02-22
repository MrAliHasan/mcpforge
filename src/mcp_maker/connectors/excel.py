"""
MCP-Maker Excel Connector — Inspect Excel (.xlsx) files.

Each sheet in the workbook becomes a table, with the first row as column headers.
"""

import os

from .base import BaseConnector, register_connector
from .utils import sanitize_name, infer_type
from ..core.schema import (
    Column,
    ColumnType,
    DataSourceSchema,
    Table,
)


# Map Python/openpyxl types to our universal types
def _infer_column_type(values: list) -> ColumnType:
    """Infer column type from sample values."""
    non_none = [v for v in values if v is not None]
    if not non_none:
        return ColumnType.STRING

    type_counts = {"int": 0, "float": 0, "bool": 0, "str": 0}
    for v in non_none[:100]:  # Sample first 100 rows
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, int):
            type_counts["int"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        else:
            type_counts["str"] += 1

    dominant = max(type_counts, key=type_counts.get)
    return {
        "int": ColumnType.INTEGER,
        "float": ColumnType.FLOAT,
        "bool": ColumnType.BOOLEAN,
        "str": ColumnType.STRING,
    }[dominant]


class ExcelConnector(BaseConnector):
    """Connector for Excel (.xlsx) files.

    Inspects all sheets, treating each sheet as a table.
    The first row is used as column headers.

    URI format: excel:///path/to/file.xlsx or just ./file.xlsx
    """

    @property
    def source_type(self) -> str:
        return "excel"

    def _get_file_path(self) -> str:
        """Extract the file path from the URI."""
        path = self.uri
        if path.startswith("excel:///"):
            path = path[len("excel:///"):]
        elif path.startswith("excel://"):
            path = path[len("excel://"):]
        return os.path.expanduser(path)

    def validate(self) -> bool:
        """Check that the Excel file exists and is readable."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install it with: pip install mcp-maker[excel]"
            )

        file_path = self._get_file_path()
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx files.")

        return True

    def inspect(self) -> DataSourceSchema:
        """Inspect the Excel file and return its schema."""
        import openpyxl

        file_path = self._get_file_path()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # First row = headers
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]

            # Skip empty sheets (no headers)
            if not any(h for h in headers):
                continue

            # Infer column types from data
            columns = []
            for col_idx, header in enumerate(headers):
                col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
                col_type = _infer_column_type(col_values)
                safe_name = sanitize_name(header)
                columns.append(Column(
                    name=safe_name,
                    type=col_type,
                    nullable=True,
                ))

            safe_sheet = sanitize_name(sheet_name)
            tables.append(Table(
                name=safe_sheet,
                columns=columns,
                row_count=len(data_rows),
                description=sheet_name,
            ))

        wb.close()

        return DataSourceSchema(
            source_type="excel",
            source_uri=self.uri,
            tables=tables,
            metadata={
                "file_path": file_path,
                "file_size_bytes": os.path.getsize(file_path),
                "sheet_count": len(tables),
            },
        )


# Register this connector
register_connector("excel", ExcelConnector)
